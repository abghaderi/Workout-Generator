"""
اپ وب ساده تولید برنامه تمرینی — نسخه Streamlit
--------------------------------------------------
اجرا:
    pip install streamlit openpyxl
    streamlit run workout_app.py

این اپ یک رابط کاربری ساده (بدون نیاز به کدنویسی) روی همان منطق
اسکریپت قبلی می‌سازد. همسر شما فقط از منوها انتخاب می‌کند و
دکمه می‌زند؛ فایل اکسل نهایی آماده دانلود می‌شود.

در نسخه واقعی، به‌جای EXERCISES هاردکد، این لیست را از Google Sheets
یا Airtable بخوانید (چند خط تغییر لازم دارد، نه بازنویسی کل اپ).
"""

import random
import io
from dataclasses import dataclass, field
from datetime import datetime

import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ---------------------------------------------------------------------------
# دیتابیس تمرین‌ها (بعداً از Google Sheets خوانده می‌شود)
# ---------------------------------------------------------------------------

EXERCISES = [
    {"name": "اسکوات با هالتر", "muscle": "پا", "sport": "بدنسازی", "level": "متوسط", "equipment": "هالتر"},
    {"name": "پرس سینه با دمبل", "muscle": "سینه", "sport": "بدنسازی", "level": "مبتدی", "equipment": "دمبل"},
    {"name": "ددلیفت رومانیایی", "muscle": "پا", "sport": "بدنسازی", "level": "پیشرفته", "equipment": "هالتر"},
    {"name": "زیربغل سیمکش", "muscle": "پشت", "sport": "بدنسازی", "level": "مبتدی", "equipment": "سیمکش"},
    {"name": "شنا سوئدی", "muscle": "سینه", "sport": "بدنسازی", "level": "مبتدی", "equipment": "بدون تجهیزات"},
    {"name": "پرس شانه دمبل", "muscle": "شانه", "sport": "بدنسازی", "level": "متوسط", "equipment": "دمبل"},
    {"name": "پلانک", "muscle": "شکم", "sport": "فانکشنال فیتنس", "level": "مبتدی", "equipment": "بدون تجهیزات"},
    {"name": "برپی", "muscle": "کل بدن", "sport": "فانکشنال فیتنس", "level": "متوسط", "equipment": "بدون تجهیزات"},
    {"name": "کتل‌بل سوئینگ", "muscle": "کل بدن", "sport": "فانکشنال فیتنس", "level": "متوسط", "equipment": "کتل‌بل"},
    {"name": "باکس جامپ", "muscle": "پا", "sport": "فانکشنال فیتنس", "level": "پیشرفته", "equipment": "باکس"},
    {"name": "ضربات جب-کراس", "muscle": "کل بدن", "sport": "بوکس", "level": "مبتدی", "equipment": "دستکش/کیسه"},
    {"name": "شدو باکسینگ", "muscle": "کل بدن", "sport": "بوکس", "level": "مبتدی", "equipment": "بدون تجهیزات"},
    {"name": "ترکیب هوک-آپرکات", "muscle": "کل بدن", "sport": "بوکس", "level": "متوسط", "equipment": "دستکش/کیسه"},
    {"name": "جامپینگ جک", "muscle": "کل بدن", "sport": "ایرفیت", "level": "مبتدی", "equipment": "بدون تجهیزات"},
    {"name": "های نیز", "muscle": "کل بدن", "sport": "ایرفیت", "level": "مبتدی", "equipment": "بدون تجهیزات"},
    {"name": "استپ تاچ ترکیبی", "muscle": "کل بدن", "sport": "ایرفیت", "level": "متوسط", "equipment": "بدون تجهیزات"},
]

LEVEL_ORDER = {"مبتدی": 1, "متوسط": 2, "پیشرفته": 3}
SPORTS = sorted({ex["sport"] for ex in EXERCISES})
LEVELS = ["مبتدی", "متوسط", "پیشرفته"]
GOALS = ["کاهش وزن", "افزایش قدرت", "استقامت", "فرم‌دهی بدن"]
EQUIPMENT_OPTIONS = sorted({ex["equipment"] for ex in EXERCISES})


@dataclass
class ClientProfile:
    name: str
    sport: str
    level: str
    goal: str
    days_per_week: int
    available_equipment: list = field(default_factory=list)


def filter_exercises(client: ClientProfile):
    result = []
    for ex in EXERCISES:
        if ex["sport"] != client.sport:
            continue
        if LEVEL_ORDER[ex["level"]] > LEVEL_ORDER[client.level]:
            continue
        if client.available_equipment and ex["equipment"] not in client.available_equipment:
            continue
        result.append(ex)
    return result


def build_weekly_program(client: ClientProfile, exercises_per_day: int = 5):
    pool = filter_exercises(client)
    if not pool:
        return None
    program = {}
    for day in range(1, client.days_per_week + 1):
        day_exercises = random.sample(pool, k=min(exercises_per_day, len(pool)))
        program[f"روز {day}"] = day_exercises
    return program


def export_to_excel_bytes(client: ClientProfile, program: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "برنامه تمرینی"
    ws.sheet_view.rightToLeft = True

    header_font = Font(name="Arial", size=14, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5233", end_color="2F5233", fill_type="solid")
    subheader_font = Font(name="Arial", size=12, bold=True, color="FFFFFF")
    subheader_fill = PatternFill(start_color="6B8E5A", end_color="6B8E5A", fill_type="solid")
    normal_font = Font(name="Arial", size=11)
    thin_border = Border(*(Side(style="thin"),) * 4)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:D1")
    ws["A1"] = f"برنامه تمرینی: {client.name}"
    ws["A1"].font = header_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = center
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:D2")
    ws["A2"] = f"رشته: {client.sport}   |   سطح: {client.level}   |   هدف: {client.goal}"
    ws["A2"].font = Font(name="Arial", size=11, italic=True)
    ws["A2"].alignment = center

    row = 4
    for day, day_exercises in program.items():
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = day
        ws[f"A{row}"].font = subheader_font
        ws[f"A{row}"].fill = subheader_fill
        ws[f"A{row}"].alignment = center
        row += 1

        headers = ["تمرین", "عضله هدف", "تجهیزات", "ست × تکرار"]
        for col, h in enumerate(headers, start=1):
            c = ws.cell(row=row, column=col, value=h)
            c.font = Font(name="Arial", bold=True)
            c.alignment = center
            c.border = thin_border
        row += 1

        for ex in day_exercises:
            sets_reps = "۴ × ۱۲" if client.goal == "کاهش وزن" else "۳ × ۸-۱۰"
            values = [ex["name"], ex["muscle"], ex["equipment"], sets_reps]
            for col, v in enumerate(values, start=1):
                c = ws.cell(row=row, column=col, value=v)
                c.font = normal_font
                c.alignment = center
                c.border = thin_border
            row += 1
        row += 1

    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 14

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()


# ---------------------------------------------------------------------------
# رابط کاربری
# ---------------------------------------------------------------------------

st.set_page_config(page_title="ساخت برنامه تمرینی", page_icon="💪", layout="centered")

st.markdown(
    "<div style='text-align: right; direction: rtl;'><h1>💪 ساخت برنامه تمرینی</h1></div>",
    unsafe_allow_html=True,
)
st.markdown(
    "<div style='text-align: right; direction: rtl;'>مشخصات مشتری را وارد کنید و برنامه به‌صورت خودکار ساخته می‌شود.</div>",
    unsafe_allow_html=True,
)
st.write("")

with st.form("client_form"):
    client_name = st.text_input("نام مشتری")
    col1, col2 = st.columns(2)
    with col1:
        sport = st.selectbox("رشته ورزشی", SPORTS)
        level = st.selectbox("سطح", LEVELS)
    with col2:
        goal = st.selectbox("هدف", GOALS)
        days = st.slider("تعداد روز تمرین در هفته", 1, 6, 3)

    equipment = st.multiselect(
        "تجهیزات موجود (خالی = همه تجهیزات)",
        EQUIPMENT_OPTIONS,
    )
    exercises_per_day = st.slider("تعداد تمرین در هر روز", 3, 8, 5)

    submitted = st.form_submit_button("🏋️ ساخت برنامه")

if submitted:
    if not client_name.strip():
        st.error("لطفاً نام مشتری را وارد کنید.")
    else:
        client = ClientProfile(
            name=client_name.strip(),
            sport=sport,
            level=level,
            goal=goal,
            days_per_week=days,
            available_equipment=equipment,
        )
        program = build_weekly_program(client, exercises_per_day=exercises_per_day)

        if program is None:
            st.error("با این ترکیب (رشته/سطح/تجهیزات)، تمرینی در دیتابیس پیدا نشد.")
        else:
            st.success(f"برنامه {client_name} با موفقیت ساخته شد ✅")

            for day, day_exercises in program.items():
                st.markdown(f"**{day}**")
                st.table(
                    [
                        {
                            "تمرین": ex["name"],
                            "عضله هدف": ex["muscle"],
                            "تجهیزات": ex["equipment"],
                        }
                        for ex in day_exercises
                    ]
                )

            excel_bytes = export_to_excel_bytes(client, program)
            filename = f"{client_name}_{datetime.now().strftime('%Y%m%d')}.xlsx"
            st.download_button(
                label="⬇️ دانلود فایل اکسل",
                data=excel_bytes,
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
